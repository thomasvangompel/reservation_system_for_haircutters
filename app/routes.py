from app.models import Employee


from email.mime.text import MIMEText
import smtplib
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Blueprint, render_template, redirect, url_for, request, session, flash, send_file
from app.models import db, User, Customer, CalendarEvent, Skill, Reservation
import pandas as pd
from io import BytesIO
from odf.opendocument import OpenDocumentSpreadsheet
from odf.style import Style, TextProperties, TableCellProperties, ParagraphProperties
from odf.table import Table, TableRow, TableCell
from odf.text import P
from flask import Blueprint
from app.forms import (
    ReservationForm,
    ProfileForm,
    SkillForm,
    RegistrationForm,
    LoginForm,
    WerknemerForm,
    CustomerForm
)
main = Blueprint('main', __name__)

@main.route('/download_agenda_excel')
def download_agenda_excel():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    user_id = session['user_id']
    reservations = Reservation.query.filter_by(user_id=user_id).order_by(Reservation.created_at.desc()).all()
    data = [
        {
            'Customer Name': r.customer_name,
            'Email': r.customer_email,
            'Service': r.skill_id,
            'Block': r.block,
            'Created At': r.created_at
        } for r in reservations
    ]
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    # Set column widths
    col_widths = [18, 28, 10, 32, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    # Add borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    # Format date column
    for cell in ws["E"]:
        if cell.row == 1:
            continue
        cell.number_format = 'YYYY-MM-DD HH:MM:SS'
        cell.alignment = Alignment(horizontal="center")
    # Save to BytesIO
    output2 = BytesIO()
    wb.save(output2)
    output2.seek(0)
    return send_file(output2, download_name='agenda.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@main.route('/download_agenda_odt')
def download_agenda_odt():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    user_id = session['user_id']
    reservations = Reservation.query.filter_by(user_id=user_id).order_by(Reservation.created_at.desc()).all()
    doc = OpenDocumentSpreadsheet()
    table = Table(name="Agenda")
    # Define bold text style for header
    from odf.style import Style, TextProperties, TableCellProperties, ParagraphProperties
    # Header cell style: bold, blue background, center, padding
    header_text_style = Style(name="HeaderTextStyle", family="paragraph")
    header_text_style.addElement(TextProperties(fontweight="bold", color="#FFFFFF", backgroundcolor="#4F81BD"))
    header_text_style.addElement(ParagraphProperties(textalign="center"))
    doc.styles.addElement(header_text_style)
    # Header row
    header_row = TableRow()
    for col in ['Customer Name', 'Email', 'Service', 'Block', 'Created At']:
        cell = TableCell()
        cell.addElement(P(text=col, stylename=header_text_style))
        header_row.addElement(cell)
    table.addElement(header_row)
    # Data rows
    for r in reservations:
        row = TableRow()
        for val in [r.customer_name, r.customer_email, r.skill_id, r.block, str(r.created_at)]:
            cell = TableCell()
            cell.addElement(P(text=str(val)))
            row.addElement(cell)
        table.addElement(row)
    doc.spreadsheet.addElement(table)
    output = BytesIO()
    doc.write(output)
    output.seek(0)
    return send_file(output, download_name='agenda.ods', as_attachment=True, mimetype='application/vnd.oasis.opendocument.spreadsheet')
@main.route('/<storename>/reservation', methods=['GET', 'POST'])
def reservation(storename):
    user = User.query.filter_by(store_name=storename).first()
    if not user:
        flash('Store not found!')
        return redirect(url_for('main.login'))
    user_id = user.id
    import datetime
    skills = Skill.query.filter_by(user_id=user_id).all()
    events = CalendarEvent.query.filter_by(user_id=user_id).all()
    # Generate available blocks for each skill
    for skill in skills:
        interval = skill.duration if skill.duration else 15  # default to 15 min if not set
        available_blocks = []
        for event in events:
            if event.start_time and event.end_time:
                def time_to_minutes(t):
                    h, m = map(int, t.split(':'))
                    return h * 60 + m
                def minutes_to_time(m):
                    h = m // 60
                    m = m % 60
                    return f"{h:02d}:{m:02d}"
                start = time_to_minutes(event.start_time)
                end = time_to_minutes(event.end_time)
                # Build a list of all possible block starts
                block_starts = list(range(start, end, interval))
                for block_start in block_starts:
                    block_end = block_start + skill.duration
                    if block_end > end:
                        continue  # Not enough time left in this event
                    # Check if all sub-blocks are free
                    is_free = True
                    for check_start in range(block_start, block_end, interval):
                        check_end = check_start + interval
                        block_str = f"{event.date} {minutes_to_time(check_start)} - {minutes_to_time(check_end)}"
                        reserved = Reservation.query.filter_by(user_id=user_id, block=block_str).first()
                        if reserved:
                            is_free = False
                            break
                    if is_free:
                        block_str = f"{event.date} {minutes_to_time(block_start)} - {minutes_to_time(block_end)}"
                        available_blocks.append(block_str)
        skill.available_times = available_blocks

    form = ReservationForm()
    if form.validate_on_submit():
        customer_name = form.customer_name.data
        customer_email = form.customer_email.data
        customer_phone = form.customer_phone.data
        skill_id = form.skill_id.data
        block = form.block.data
        # Add customer if not already present
        existing_customer = Customer.query.filter_by(name=customer_name, user_id=user_id).first()
        if not existing_customer:
            new_customer = Customer(name=customer_name, phone=customer_phone, email=customer_email, user_id=user_id)
            db.session.add(new_customer)
        elif existing_customer:
            # Update phone or email if not set or changed
            if not existing_customer.phone or existing_customer.phone != customer_phone:
                existing_customer.phone = customer_phone
            if not existing_customer.email or existing_customer.email != customer_email:
                existing_customer.email = customer_email
        # Save reservation
        new_res = Reservation(customer_name=customer_name, customer_email=customer_email, skill_id=skill_id, block=block, user_id=user_id)
        db.session.add(new_res)
        # Remove reserved block from CalendarEvent
        block_date, times = block.split(' ', 1)
        start_time, end_time = times.split(' - ')
        # Find and delete the reserved event
        reserved_event = CalendarEvent.query.filter_by(user_id=user_id, date=block_date, start_time=start_time, end_time=end_time).first()
        if reserved_event:
            db.session.delete(reserved_event)
        # Remove overlapping blocks if skill duration is longer than the block
        skill = Skill.query.get(skill_id)
        if skill and skill.duration:
            # Convert times to minutes
            def time_to_minutes(t):
                h, m = map(int, t.split(':'))
                return h * 60 + m
            block_start = time_to_minutes(start_time)
            block_end = block_start + skill.duration
            overlapping_events = CalendarEvent.query.filter_by(user_id=user_id, date=block_date).all()
            for event in overlapping_events:
                ev_start = time_to_minutes(event.start_time)
                ev_end = time_to_minutes(event.end_time)
                # If event overlaps with reserved block
                if (ev_start < block_end and ev_end > block_start):
                    db.session.delete(event)
        db.session.commit()
        # Send email to customer
        send_reservation_email(customer_email, block)
        flash(f'Reservation for {customer_name} at {block} has been saved and an email has been sent to {customer_email}!')
        return redirect(url_for('main.thankyou'))
    return render_template('reservation.html', skills=skills, form=form)

def send_reservation_email(to_email, block):
    # Gmail SMTP configuration
    smtp_server = 'smtp.gmail.com'
    smtp_port = 587
    smtp_user = 'yourgmail@gmail.com'  # Replace with your Gmail address
    smtp_pass = 'your_app_password'    # Use an App Password, not your main Gmail password
    msg = MIMEText(f'Your reservation is confirmed for {block}.')
    msg['Subject'] = 'Reservation Confirmation'
    msg['From'] = smtp_user
    msg['To'] = to_email
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, [to_email], msg.as_string())
    except Exception as e:
        print('Email error:', e)

@main.route('/agenda')
def agenda():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    user_id = session['user_id']
    reservations = Reservation.query.filter_by(user_id=user_id).order_by(Reservation.created_at.desc()).all()
    return render_template('agenda.html', reservations=reservations)

@main.route('/bulk_delete_availability', methods=['POST'])
def bulk_delete_availability():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    ids = request.form.getlist('delete_ids')
    if ids:
        for event_id in ids:
            event = CalendarEvent.query.get(event_id)
            if event and event.user_id == session['user_id']:
                db.session.delete(event)
        db.session.commit()
        flash('Selected availabilities deleted!')
    else:
        flash('No availabilities selected!')
    return redirect(url_for('main.calendar'))

@main.route('/')
def home():
    return render_template('home.html')

@main.route('/register', methods=['GET', 'POST'])
def register():
    form = RegistrationForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        if User.query.filter_by(username=username).first():
            flash('User already exists!')
            return redirect(url_for('main.register'))
        hashed_pw = generate_password_hash(password)
        new_user = User(username=username, password=hashed_pw)
        db.session.add(new_user)
        db.session.commit()
        flash('Registration successful! Please log in.')
        return redirect(url_for('main.login'))
    return render_template('register.html', form=form)

@main.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['username'] = user.username
            return redirect(url_for('main.dashboard'))
        flash('Invalid credentials!')
    return render_template('login.html', form=form)

@main.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    return render_template('dashboard.html', username=session['username'])

@main.route('/customers', methods=['GET', 'POST'])
def customers():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    user_id = session['user_id']
    form = CustomerForm()
    if form.validate_on_submit():
        new_customer = Customer(
            name=form.name.data,
            phone=form.phone.data,
            email=form.email.data,
            user_id=user_id
        )
        db.session.add(new_customer)
        db.session.commit()
        flash('Customer added!')
    customer_list = Customer.query.filter_by(user_id=user_id).all()
    return render_template('customers.html', customers=customer_list, form=form)

@main.route('/calendar', methods=['GET', 'POST'])
def calendar():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    user_id = session['user_id']
    page = request.args.get('page', 1, type=int)
    per_page = 5
    if request.method == 'POST':
        title = request.form['title']
        date = request.form['date']
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')
        selected_days = request.form.get('selected_days')
        if selected_days:
            import datetime
            days = selected_days.split(',')
            base_date = datetime.datetime.strptime(date, "%Y-%m-%d")
            weekday_map = {
                'Monday': 0,
                'Tuesday': 1,
                'Wednesday': 2,
                'Thursday': 3,
                'Friday': 4,
                'Saturday': 5,
                'Sunday': 6
            }
            for day in days:
                # Find next date for each selected weekday
                day_num = weekday_map[day]
                delta_days = (day_num - base_date.weekday()) % 7
                target_date = base_date + datetime.timedelta(days=delta_days)
                new_event = CalendarEvent(title=title, date=target_date.strftime("%Y-%m-%d"), start_time=start_time, end_time=end_time, user_id=user_id)
                db.session.add(new_event)
            db.session.commit()
            flash('Availability added for selected days!')
        else:
            new_event = CalendarEvent(title=title, date=date, start_time=start_time, end_time=end_time, user_id=user_id)
            db.session.add(new_event)
            db.session.commit()
            flash('Availability added!')
    events = CalendarEvent.query.filter_by(user_id=user_id).order_by(CalendarEvent.date, CalendarEvent.start_time).paginate(page=page, per_page=per_page)
    return render_template('calendar.html', events=events)

@main.route('/delete_availability/<int:event_id>', methods=['POST'])
def delete_availability(event_id):
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    event = CalendarEvent.query.get_or_404(event_id)
    if event.user_id != session['user_id']:
        flash('Unauthorized!')
        return redirect(url_for('main.calendar'))
    db.session.delete(event)
    db.session.commit()
    flash('Availability deleted!')
    return redirect(url_for('main.calendar'))

@main.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('username', None)
    return redirect(url_for('main.login'))


@main.route('/profile', methods=['GET', 'POST'])
def profile():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    user = User.query.get(session['user_id'])
    form = ProfileForm(obj=user)
    if form.validate_on_submit():
        form.populate_obj(user)
    if request.method == 'POST':
        user.name = request.form.get('name')
        user.email = request.form.get('email')
        user.store_name = request.form.get('store_name')
        user.street = request.form.get('street')
        user.postal_code = request.form.get('postal_code')
        user.city = request.form.get('city')
        user.country = request.form.get('country')
        db.session.commit()
        flash('Profile updated!')
        return redirect(url_for('main.dashboard'))
    return render_template('profile.html', form=form, user=user)

@main.route('/skills', methods=['GET', 'POST'])
def skills():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    user_id = session['user_id']
    form = SkillForm()
    if form.validate_on_submit():
        image = form.image.data
        image_url = None
        if image and hasattr(image, 'filename') and image.filename:
            image_path = f'static/uploads/{image.filename}'
            image.save(f'app/{image_path}')
            image_url = '/' + image_path
        new_skill = Skill(
            gender=form.gender.data,
            type=form.type.data,
            price=form.price.data,
            duration=form.duration.data,
            image_url=image_url
        )
        db.session.add(new_skill)
        db.session.commit()
        flash('Skill added!')
    page = request.args.get('page', 1, type=int)
    skills = Skill.query.filter_by(user_id=user_id).paginate(page=page, per_page=5)
    return render_template('skills.html', skills=skills, form=form)
    page = request.args.get('page', 1, type=int)
    per_page = 3
    skills = Skill.query.filter_by(user_id=user_id).paginate(page=page, per_page=per_page)
    return render_template('skills.html', skills=skills)

@main.route('/delete_skill/<int:skill_id>', methods=['POST'])
def delete_skill(skill_id):
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    skill = Skill.query.get_or_404(skill_id)
    if skill.user_id != session['user_id']:
        flash('Unauthorized!')
        return redirect(url_for('main.skills'))
    db.session.delete(skill)
    db.session.commit()
    flash('Skill deleted!')
    return redirect(url_for('main.skills'))

@main.route('/thankyou')
def thankyou():
    return render_template('thankyou.html')


@main.route('/werknemers', methods=['GET', 'POST'])
def werknemers():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))
    user_id = session['user_id']
    form = WerknemerForm()
    if form.validate_on_submit():
        new_employee = Employee(
            name=form.name.data,
            email=form.email.data,
            address=form.address.data,
            phone=form.phone.data,
            age=form.age.data,
            gender=form.gender.data,
            user_id=user_id
        )
        db.session.add(new_employee)
        db.session.commit()
        flash(f"Employee {form.name.data} added!")
        return redirect(url_for('main.werknemers'))
    employees = Employee.query.filter_by(user_id=user_id).all()
    return render_template('werknemers.html', form=form, employees=employees)